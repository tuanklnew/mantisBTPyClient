import ClassBase
import APIFunction
import GlobleConstant
from argparse import ArgumentParser
import openpyxl
import ParseXmlReport
from urllib3 import disable_warnings
from requests import Session
from requests import exceptions as req_exception
from zeep import exceptions as zeep_exception
from zeep import Client
from zeep.cache import SqliteCache
from zeep.transports import Transport
from openpyxl.utils import get_column_letter
from datetime import datetime
from prettytable import PrettyTable


def main():
    cli_parser = ArgumentParser(description="Mantis Bug Tracker Client CLI-based")
    subparsers = cli_parser.add_subparsers(help="sub-command help", dest='command')

    # Create sub command for list project
    list_prj_parser = subparsers.add_parser("list-project", help="List projects in database")
    list_prj_parser.add_argument('-w', '--wsdl', required=True, help="wsdl link of Mantis Bug Tracker", type=str)
    list_prj_parser.add_argument('-u', '--username', required=True, help="Username of Mantis Bug Tracker account", type=str)
    list_prj_parser.add_argument('-p', '--password', required=True, help="Password of Mantis Bug Tracker account", type=str)
    
    # Create sub command for list issue
    list_iss_parser = subparsers.add_parser("list-issue", help="List issues in database")
    list_iss_parser.add_argument('-w', '--wsdl', required=True, help="wsdl link of Mantis Bug Tracker", type=str)
    list_iss_parser.add_argument('-u', '--username', required=True, help="Username of Mantis Bug Tracker account", type=str)
    list_iss_parser.add_argument('-p', '--password', required=True, help="Password of Mantis Bug Tracker account", type=str)
    list_iss_parser.add_argument('-i', '--pid', required=True, help="ID of project that contains information", type=int)

    # Create sub command for get project
    get_prj_parser = subparsers.add_parser("get-project", help="Get project information")
    get_prj_parser.add_argument('-w', '--wsdl', required=True, help="wsdl link of Mantis Bug Tracker", type=str)
    get_prj_parser.add_argument('-u', '--username', required=True, help="Username of Mantis Bug Tracker account", type=str)
    get_prj_parser.add_argument('-p', '--password', required=True, help="Password of Mantis Bug Tracker account", type=str)
    get_prj_parser.add_argument('-i', '--pid', required=True, help="ID of project that contains information", type=int)

    # Create sub command for get issue
    get_iss_parser = subparsers.add_parser("get-issue", help="Get issue information")
    get_iss_parser.add_argument('-w', '--wsdl', required=True, help="wsdl link of Mantis Bug Tracker", type=str)
    get_iss_parser.add_argument('-u', '--username', required=True, help="Username of Mantis Bug Tracker account", type=str)
    get_iss_parser.add_argument('-p', '--password', required=True, help="Password of Mantis Bug Tracker account", type=str)
    get_iss_parser.add_argument('-s', '--iid', required=True, help="ID of issue", type=int)

    # Create sub command for adding issues
    add_parser = subparsers.add_parser("add", help="Add issues to database")
    add_parser.add_argument('-w', '--wsdl', required=True, help="wsdl link of Mantis Bug Tracker", type=str)
    add_parser.add_argument('-u', '--username', required=True, help="Username of Mantis Bug Tracker account", type=str)
    add_parser.add_argument('-p', '--password', required=True, help="Password of Mantis Bug Tracker account", type=str)
    add_grp_me_parser = add_parser.add_mutually_exclusive_group(required=True)
    add_grp_me_parser.add_argument('-e', '--excel', help="Location of Excel file that contains issues to add", type=str)
    add_grp_me_parser.add_argument('-x', '--xml', help="Location of xml file that contains issues to add", type=str)
    add_parser.add_argument('-i', '--pid', help="ID of project that contains information", type=int)
    add_parser.add_argument('-a', '--attachment', nargs='+', help="Attachment locations to upload", type=str)

    args = cli_parser.parse_args()
    # Create soap client with cache
    disable_warnings()      # disable ssl verify warning
    session = Session()
    session.verify = False
    transport = Transport(cache=SqliteCache(), session=session, )

    # Check connection to wsdl
    try:
        soap_client = Client(wsdl=args.wsdl, transport=transport)
    except req_exception.ConnectionError as error:
        print("[!] Connection Error:\t{}".format(error))
        return -1
    factory = soap_client.type_factory('ns0')
    if args.command == 'list-project':
        list_project(soap_client, args)
    elif args.command == 'list-issue':
        list_issue(soap_client, factory, args)
    elif args.command == 'get-project':
        get_project(soap_client, args)
    elif args.command == 'get-issue':
        get_issue(soap_client, args)
    elif args.command == 'add' and args.excel:
        if args.pid and args.attachment:
            print('[i] PID and attachment argument is only for xml file. It will be ignore')
        elif args.pid:
            print('[i] PID argument is only for xml file. It will be ignore')
        elif args.attachment:
            print('[i] Attachment argument is only for xml file. It will be ignore')
        add_issue_excel(soap_client, factory, args)
    elif args.command == 'add' and args.xml:
        add_issue_xml(soap_client, factory, args)


def list_project(soap_client, args):
    table = PrettyTable(['ID', 'Project Name'])
    try:
        projects_arr = APIFunction.mc_projects_get_user_accessible(soap_client, username=args.username, password=args.password)
    except zeep_exception.Fault:
        print('[!] Login Failed. Permission deny')
        return 0
    for project in projects_arr:
        table.add_row([project.id, project.name])
    print(table)


def list_issue(soap_client, factory, args=None):
    try:
        issues = APIFunction.mc_filter_search_issues(soap_client, filter=ClassBase.FilterSearchData(project_id=[
            APIFunction.mc_project_get_id_from_name(soap_client, username=args.username, password=args.password,
                                                    project_name=args.pid)]).parse_to_zeep_type(factory),
                                                     username=args.username, password=args.password)
    except zeep_exception.Fault:
        print('[!] Login Failed. Permission deny')
        return 0
    table = PrettyTable(['ID', 'IP Address', 'Name of Scan', 'Times of Scan', 'Date of Scan', 'Risk Level'])
    for issue in issues:
        ip_addr = None
        scan_name = None
        times_scan = None
        date_of_scan = None
        for custom_field in issue.custom_fields:
            if custom_field.field.name.lower() == 'ip address':
                ip_addr = custom_field.value
                continue
            if custom_field.field.name.lower() == 'name of scan':
                scan_name = custom_field.value
                continue
            if custom_field.field.name.lower() == 'times of scan':
                times_scan = custom_field.value
                continue
            if custom_field.field.name.lower() == 'date of scan':
                date_of_scan = custom_field.value
                continue
        table.add_row([issue.id, ip_addr, scan_name, times_scan, date_of_scan, issue.priority.name])
    print(table)


def get_project(soap_client, args=None):
    try:
        projects_arr = APIFunction.mc_projects_get_user_accessible(soap_client, username=args.username,
                                                               password=args.password)
    except zeep_exception.Fault:
        print('[!] Login Failed. Permission deny')
        return 0
    table = PrettyTable(
        ['ID', 'Project Name', 'Status', 'Enabled', 'View State', 'Access', 'Description', 'Sub-projects'])
    project_match = None
    for project in projects_arr:
        if str(project.id) == str(args.pid):
            project_match = project
            sub_project_name_arr = []
            for sub_project in project.subprojects:
                sub_project_name_arr.append(sub_project.name)
            table.add_row([project.id, project.name, project.status.name, project.enabled, project.view_state.name, project.access_min.name, project.description, sub_project_name_arr])
            break
    if not project_match:
        print('[!] MantisBT does not have PID {}'.format(args.pid))
        return 0
    print(table)


def get_issue(soap_client, args):
    #table = PrettyTable(['Issue ID', 'View State', 'Project', 'Priority', 'Severity', 'Status', 'Description', 'IP Address', 'Name of Scan', 'Times of Scan', 'Date of Scan',])
    try:
        issue = APIFunction.mc_issue_get(soap_client, username=args.username, password=args.password, issue_id=args.iid)
    except zeep_exception.Fault as err:
        if err == 'Access denied':
            print('[!] Login Failed. Permission deny')
        else:
            print('[!] {}'.format(err))
        return 0
    print(issue)


def add_issue_excel(soap_client, factory, args=None):
    if args.excel:
        # Open Excel file to process
        try:
            input_file_excel = openpyxl.load_workbook(filename=args.excel)
        except FileNotFoundError as err:
            print(err)
            return 0

        input_sheet = input_file_excel.active

        # Get number of column
        num_col = input_sheet.max_column

        # Get number of rows
        num_row = input_sheet.max_row

        # Global Const
        view_state_glob_const = GlobleConstant.ViewStateGlobConst()
        priority_glob_const = GlobleConstant.PriorityGlobConst()
        severity_glob_const = GlobleConstant.SeverityGlobConst()
        status_glob_const = GlobleConstant.StatusGlobConst()
        reproducibility_glob_const = GlobleConstant.ReproducibilityGlobConst()
        resolution_glob_const = GlobleConstant.ResolutionsGlobConst()

        # Get info
        num_issue_added = 0
        for row in range(2, num_row + 1):
            # Clear Var to store parsed info
            ip_addr = ''
            name_of_scan = ''
            date_of_scan = None
            view_state = None
            last_update = None
            project = None
            category = None
            priority = None
            severity = None
            status = None
            summary = None
            platform = ''
            os = ''
            os_build = ''
            reproducibility = ''
            date_submitted = ''
            resolution = None
            description = ''
            steps_to_reproduce = ''
            attachments = []
            additional_information = ''
            flag_skip = False
            id_cell_str = ''

            # Loop to process each column
            for col in range(1, num_col + 1):
                cell_str = get_column_letter(col) + str(row)
                ref_cell_str = get_column_letter(col) + str(1)

                # ID field in file must be only written by programs. Issue will be skipped if it already has ID
                if input_sheet[ref_cell_str].value.lower() == 'id':
                    if input_sheet[cell_str].value is not None:
                        print("[+] ID issue: {} has value in file. Program will skip this issue".format(
                            input_sheet[cell_str].value))
                        flag_skip = True
                        break
                    else:
                        id_cell_str = cell_str

                # Get ip address
                elif input_sheet[ref_cell_str].value.lower() == 'ip_addr':
                    ip_addr = input_sheet[cell_str].value

                # Get ip address
                elif input_sheet[ref_cell_str].value.lower() == 'name_of_scan':
                    name_of_scan = input_sheet[cell_str].value

                # Get date_of_scan
                elif input_sheet[ref_cell_str].value.lower() == 'date_of_scan':
                    if input_sheet[cell_str].value == "":
                        date_of_scan = str(datetime.today())
                    else:
                        date_of_scan = str(input_sheet[cell_str].value)

                # Get view_state
                elif input_sheet[ref_cell_str].value.lower() == 'view_state':
                    view_state = view_state_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get last_update
                elif input_sheet[ref_cell_str].value.lower() == 'last_update':
                    if input_sheet[cell_str].value is not None:
                        last_update = input_sheet[cell_str].value
                    else:
                        last_update = datetime.today()

                # Get Project
                elif input_sheet[ref_cell_str].value.lower() == 'project':
                    project_name = input_sheet[cell_str].value
                    try:
                        project_id = APIFunction.mc_project_get_id_from_name(soap_client, username=args.username,
                                                                         password=args.password,
                                                                         project_name=project_name)
                    except zeep_exception.Fault:
                        print('[!] Login Failed. Permission deny')
                        return 0
                    project = ClassBase.ObjectRef(id=project_id, name=project_name)
                # Get category
                elif input_sheet[ref_cell_str].value.lower() == 'category':
                    if input_sheet[cell_str].value is not None:
                        category = input_sheet[cell_str].value
                    else:
                        category = 'General'

                # Get priority
                elif input_sheet[ref_cell_str].value.lower() == 'priority':
                    priority = priority_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get severity
                elif input_sheet[ref_cell_str].value.lower() == 'severity':
                    severity = severity_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get status
                elif input_sheet[ref_cell_str].value.lower() == 'status':
                    status = status_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get summary
                elif input_sheet[ref_cell_str].value.lower() == 'summary':
                    summary = input_sheet[cell_str].value

                # Get platform
                elif input_sheet[ref_cell_str].value.lower() == 'platform':
                    platform = input_sheet[cell_str].value

                # Get os
                elif input_sheet[ref_cell_str].value.lower() == 'os':
                    os = input_sheet[cell_str].value

                # Get os_build
                elif input_sheet[ref_cell_str].value.lower() == 'os_build':
                    os_build = input_sheet[cell_str].value

                # Get reproducibility
                elif input_sheet[ref_cell_str].value.lower() == 'reproducibility':
                    reproducibility = reproducibility_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get date_submitted
                elif input_sheet[ref_cell_str].value.lower() == 'date_submitted':
                    if input_sheet[cell_str].value is not None:
                        date_submitted = input_sheet[cell_str].value
                    else:
                        date_submitted = datetime.today()

                # Get resolution
                elif input_sheet[ref_cell_str].value.lower() == 'resolution':
                    resolution = resolution_glob_const.__getattribute__(input_sheet[cell_str].value)

                # Get description
                elif input_sheet[ref_cell_str].value.lower() == 'description':
                    description = input_sheet[cell_str].value

                # Get steps_to_reproduce
                elif input_sheet[ref_cell_str].value.lower() == 'steps_to_reproduce':
                    steps_to_reproduce = input_sheet[cell_str].value

                # Get attachments
                elif input_sheet[ref_cell_str].value.lower() == 'attachments':
                    attachments = str(input_sheet[cell_str].value).split(", ")
                # Get additional_information
                elif input_sheet[ref_cell_str].value.lower() == 'additional_information':
                    additional_information = input_sheet[cell_str].value
            if not flag_skip:
                ret_id_issue = APIFunction.add_issue(soap_client,
                                                     factory=factory,
                                                     username=args.username,
                                                     password=args.password,
                                                     ip_addr=ip_addr,
                                                     name_of_scan=name_of_scan,
                                                     date_of_scan=date_of_scan,
                                                     view_state=view_state,
                                                     last_updated=last_update,
                                                     project=project,
                                                     category=category,
                                                     priority=priority,
                                                     severity=severity,
                                                     status=status,
                                                     summary=summary,
                                                     platform=platform,
                                                     os=os,
                                                     os_build=os_build,
                                                     reproducibility=reproducibility,
                                                     date_submitted=date_submitted,
                                                     resolution=resolution,
                                                     description=description,
                                                     steps_to_reproduce=steps_to_reproduce,
                                                     attachments=attachments,
                                                     additional_information=additional_information,
                                                     due_date=datetime.today()
                                                     )
                if ret_id_issue[0] == -1:
                    print("[+] This issue is in Mantis BT already")
                elif ret_id_issue[0] == -2:
                    print("[+] Err: incorrect custom fields definition")
                    return 0
                elif ret_id_issue[0] == -3:
                    print('[!] Login Failed. Permission deny')
                    return 0
                else:
                    print("[+] Add issue id {} to Mantis BT".format(ret_id_issue[0]))
                    input_sheet[id_cell_str].value = ret_id_issue[0]
                    num_issue_added = num_issue_added + 1
                    for ret_att in ret_id_issue[1]:
                        if ret_att[1] == -1:
                            print("\t[Err]: {} can't be read".format(ret_att[0]))
                        else:
                            print("\t[-]: {} is uploaded".format(ret_att[0]))
        print("[TOTAL] {} issues of {} issues added".format(num_issue_added, num_row - 1))
        try:
            input_file_excel.save(args.excel)
        except PermissionError as err:
            print(str(err) + " - Can not save issue IDs")


def add_issue_xml(soap_client, factory, args=None):
    # if pid is not set. Then Project id is set to default value
    prj_arr = APIFunction.mc_projects_get_user_accessible(soap_client, args.username, args.password)
    project = None
    if args.pid:
        for prj in prj_arr:
            if prj.id == args.pid:
                project = ClassBase.ObjectRef(id=prj.id, name=prj.name)
        if not project:
            print("[!] Project ID is not in database")
            return
    else:
        project = ClassBase.ObjectRef(id=prj_arr[0].id, name=prj_arr[0].name)


    # set value
    view_state = GlobleConstant.ViewStateGlobConst.public
    last_update = datetime.today()
    status = GlobleConstant.StatusGlobConst.new
    reproducibility = GlobleConstant.ReproducibilityGlobConst.always
    date_submitted = datetime.today()
    due_date = None
    category = 'General'
    if args.attachment:
        attachments = args.attachment
    else:
        attachments = []
    severity = GlobleConstant.SeverityGlobConst.major

    num_issue_added = 0
    # Parse XML file report
    try:
        vulns = ParseXmlReport.parse_xml_host_data(args.xml)
    except FileNotFoundError as err:
        print(err)
        return
    for vuln in vulns:
        ip_addr = vuln.host_info.IPAddress
        name_of_scan = vuln.scan_info.ScanName
        date_of_scan = vuln.scan_info.StartTime
        os = vuln.host_info.OSName
        description = vuln.Description
        summary = vuln.VulnName
        if int(vuln.Risk) == 0:
            priority = GlobleConstant.PriorityGlobConst.none
        elif int(vuln.Risk) in range(1, 4):
            priority = GlobleConstant.PriorityGlobConst.low
        elif int(vuln.Risk) in range(4, 7):
            priority = GlobleConstant.PriorityGlobConst.normal
        else:
            priority = GlobleConstant.PriorityGlobConst.high
        ret_id_issue = APIFunction.add_issue(soap_client,
                                             factory=factory,
                                             username=args.username,
                                             password=args.password,
                                             ip_addr=ip_addr,
                                             name_of_scan=name_of_scan,
                                             date_of_scan=date_of_scan,
                                             view_state=view_state,
                                             last_updated=last_update,
                                             project=project,
                                             category=category,
                                             priority=priority,
                                             severity=severity,
                                             status=status,
                                             summary=summary,
                                             platform=None,
                                             os=os,
                                             os_build=None,
                                             reproducibility=reproducibility,
                                             date_submitted=date_submitted,
                                             resolution=None,
                                             description=description,
                                             steps_to_reproduce=None,
                                             attachments=attachments,
                                             additional_information=None,
                                             due_date=due_date
                                             )
        if ret_id_issue[0] == -1:
            print("[!] This issue is in Mantis BT already")
        elif ret_id_issue[0] == -2:
            print("[!] Err: incorrect custom fields definition")
            return 0
        elif ret_id_issue[0] == -3:
            print('[!] Login Failed. Permission deny')
            return 0
        else:
            print("[+] Add issue id {} to Mantis BT".format(ret_id_issue[0]))
            num_issue_added = num_issue_added + 1
            for ret_att in ret_id_issue[1]:
                if ret_att[1] == -1:
                    print("\t[Err]: {} can't be read".format(ret_att[0]))
                else:
                    print("\t[-]: {} is uploaded".format(ret_att[0]))
    print("[TOTAL] {} issues of {} issues added".format(num_issue_added, len(vulns)))


if __name__ == '__main__':
    print('  __  __             _   _       ____ _____    ____ _ _            _    ')
    print(' |  \/  | __ _ _ __ | |_(_)___  | __ |_   _|  / ___| (_) ___ _ __ | |_  ')
    print(' | |\/| |/ _` | \'_ \| __| / __| |  _ \ | |   | |   | | |/ _ | \'_ \| __| ')
    print(' | |  | | (_| | | | | |_| \__ \ | |_) || |   | |___| | |  __| | | | |_  ')
    print(' |_|  |_|\__,_|_| |_|\__|_|___/ |____/ |_|    \____|_|_|\___|_| |_|\__| ')
    print('\n\n')
    main()




